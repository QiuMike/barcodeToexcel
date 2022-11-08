import barcode
import sys
from barcode.writer import ImageWriter
from openpyxl import load_workbook,Workbook
from openpyxl.drawing.image import Image
from flask import send_file, send_from_directory
from urllib.parse import quote
import os
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.cell.cell import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils.units import cm_to_EMU
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D

def getcode128(message):
    fullname=""
    try:
        options = {"module_height":2,"text_distance":2,"font_size":5, "quiet_zone":0.5,"dpi":300,"format":"JPEG"}

        EAN = barcode.get_barcode_class('code128')

        ean=EAN(message,writer=ImageWriter())

        fullname = ean.save(message,options)
    except Exception as e:
        print("Exception:",str(e))

    return fullname

def get_image(sheet,start_row, start_col, hieght, img):

    col_letter = get_column_letter(start_col)

    sheet.column_dimensions[col_letter].width = (img.width/96)*2.54*4.374 + 2.5
    width = sheet.column_dimensions[col_letter].width
    _from = AnchorMarker(start_col-1, 0, start_row-1, 0)
    _to = AnchorMarker(start_col, 0, start_row, 0)
    img.anchor = TwoCellAnchor('twoCell', _from, _to)

    return img

def xlsx_process(xlsfilename,src,dest):

    m_column = dest

    start_line = 2

    file_response={}
    try:
        wb = load_workbook(xlsfilename)
        sheet = wb.active
        for l in range(start_line,sheet.max_row+1):
            # 'A' value is 65, A is column 1
            message = sheet.cell(column=ord(src)-64,row=l).value
            if message == "" or message== None:
                continue
            filename = getcode128(str(message))
            img = Image(filename)
            # 设置行高列宽，行高是磅 1cm = 28.6磅
            #sheet.row_dimensions[l].height = img.height*2.54*28.6/96
            #print("sheet.column_dimensions[m_column].width:",sheet.column_dimensions[m_column].width)
            #sheet.column_dimensions[m_column].width = (img.width/96)*2.54*4.374 +2.5
            #print("sheet.column_dimensions[m_column].width:",sheet.column_dimensions[m_column].width)
            #sheet.column_dimensions['D'].width = 20
            # 实际计算出 250像素结果28.93 但是实际上却要将近31，此处直接写了
            # 4.37 是列宽的算数乘数,同时设置的结果与实际结果竟然还差0.62，也就是到excel查看列宽为31，此处得设置为31.62
            #sheet.column_dimensions[m_column].width = 31.62
            #sheet.add_image(img,m_column+str(l))
            if sheet.row_dimensions[l].height == None:
                sheet.row_dimensions[l].height = img.height*2.54*28.6/96
            img = get_image(sheet,l, ord(m_column) - 64, sheet.row_dimensions[l].height, img)
            sheet.add_image(img)

        wb.save(xlsfilename)

        for jpegfile in os.listdir(os.getcwd()):
            if 'jpeg' in jpegfile:
                os.remove(jpegfile)

        file_response = send_from_directory('./', xlsfilename, as_attachment=True)
        # 解决中文文件名问题
        file_response.headers["Content-Disposition"] = "attachment; filename=.xlsx; filename*=utf-8''{}".format(quote(xlsfilename))

        os.remove(os.getcwd()+"/"+xlsfilename)

    except Exception as e:
        print("error:",str(e))
        return {"error":str(e)}

    return file_response
